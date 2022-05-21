import os
from posixpath import dirname
import openpyxl

wb = openpyxl.load_workbook('Log of Resources Available.xlsx')
i=1
j=1
def getListOfFiles(dirName):
    # create a list of file and sub directories 
    # names in the given directory 
    listOfFile = os.listdir(dirName)
    allFiles = list()
    # X= list()
    # Iterate over all the entries
    for entry in listOfFile:
        # Create full path
        fullPath = os.path.join(dirName, entry)
        # If entry is a directory then get the list of files in this directory 
        if os.path.isdir(fullPath):
            allFiles = allFiles + getListOfFiles(fullPath)
        else:
            # Y = fullPath.split("/")
            # print("File Name: " + Y[-1])
            # X.append(Y[-1])
            allFiles.append(fullPath)
                
    return allFiles

dirName="/media/zjaweds/zjaweds/Linux2PDrive/Learning Resources"
listOfFiles = getListOfFiles(dirName)

for item in listOfFiles:
    name =item.split("/")[-1]
    print("File Name: " + name)
    print("Full Path: "+ item)
    print(" ")
    worksheet = wb.active
    # worksheet.cell(row=0,column=0).value = "Name"
    # worksheet.cell(row=0,column=1).value = "Path"
    worksheet.cell(row=i,column=j).value = name
    worksheet.cell(row=i,column=j+1).value = item
    i+=1    
wb.save('Log of Resources Available.xlsx')